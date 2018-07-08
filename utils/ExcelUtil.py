# encoding=utf-8
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell


class ParseExcel(object):
    """Excel读取工具类"""

    def __init__(self, excel_path):
        """初始化函数"""
        # 将要读取的excel加载到内存
        self.wb = load_workbook(excel_path, read_only=True)
        # # 通过工作表名称获取一个工作表对象
        # self.sheet = self.wb.get_sheet_by_name(sheet_name)
        # # 获取工作表中存在数据的区域的最大行号
        # self.maxRowNum = self.sheet.max_row

    def get_data_from_sheet(self, sheet_index):
        """读取sheet中全部数据，返回datalist"""
        # ws = self.wb.get_sheet_by_index(sheet_index)
        ws = self.wb[self.wb.sheetnames[sheet_index]]
        # 用于存放从工作表中读取出来的数据
        sheet_data_list = []
        # 遍历Excel里内容
        for line_num, row in enumerate(ws.rows):
            if not isinstance(row[0], EmptyCell):  # 有遇到都是None的空行
                row_data_list = map(lambda it: it.value, row)  # 读取每行数据
                sheet_data_list.append(row_data_list)
        # 将获取工作表中的所有数据的迭代对象返回
        return sheet_data_list[1:]    # 第1行标题行跳过

    def get_cell_values_by_cell_index(self, sheet_index, cell_index):
        """从excel中通过行号获取该行所有数据"""
        # ws = wb.get_sheet_by_name(wb.get_sheet_names()[sheet_index])
        ws = self.wb[self.wb.sheetnames[sheet_index]]
        testcase_cn_name_list = map(lambda row: row[cell_index].value, ws.rows)
        del testcase_cn_name_list[0]  # 删除第1行数据(列名不需要)
        return testcase_cn_name_list

    def over(self):
        """函数关闭"""
        self.wb.close()


if __name__ == '__main__':
    excelPath = u'./unit_excel.xlsx'
    pe = ParseExcel(excelPath)
    data_list = pe.get_data_from_sheet(0)
    for line in data_list:
        for value in line:
            print(value,)
