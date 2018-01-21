# coding:utf-8
"""接口测试程序

数据源：Excel
测试框架：unittest
测试报告：HTMLTestRunner
"""

import HTMLTestRunner
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from utils.request_util import RequestUtil
import utils.response_util as resp_util
import unittest
import os
import traceback
import json


class ExcelUnitTest(unittest.TestCase):
    """存接口的excel文件"""

    @classmethod
    def setUpClass(cls):
        """suite的执行前方法：创建断言的值，创建工具类实例"""
        cls.PASS = "Pass"
        cls.FAIL = "Fail"
        cls.req_util = RequestUtil(5)

    def setUp(self):
        """每个testcase的执行前方法：pass"""
        pass

    def tearDown(self):
        """每个testcase的执行后方法：pass"""
        pass

    @classmethod
    def tearDownClass(cls):
        """suite的执行后方法：pass"""
        pass

    # --------------------------------------------------------------------------------
    def set_params(self, row_value_list):
        """存储入參"""
        def convert(row_value):
            if isinstance(row_value, unicode):
                return row_value.encode("utf-8")
            if isinstance(row_value, long):
                return int(row_value)
        return map(convert, row_value_list)

    def http_request(self, request_method, url, url_params):
        """请求接口"""
        request_method_dict = {
            "GET": self.req_util.get(url, url_params),
            "POST": self.req_util.post(url, url_params),
        }
        try:
            return request_method_dict[request_method.upper()]
        except KeyError:
            error_text = "request_method输入类型有误！无此http请求方法。"
            raise TypeError(error_text)

    # --------------------------------------------------------------------------------
    def check_resp(self, resp_json_obj, check_point_type, node_path, expected_result):
        """断言调度函数"""
        check_type_dict = {
            "check_json_node": self.check_json_node,
        }
        try:
            return check_type_dict[check_point_type](
                resp_json_obj, node_path, expected_result,
            )
        except KeyError:
            error_text = "check_point_type输入类型有误！无此check方法。"
            raise TypeError(error_text)

    def check_json_node(self, resp_json_obj, node_path, expected_result):
        """断言函数之一：检查json节点的值是不是等于预期结果"""
        json_node_obj = resp_util.get_json_node(resp_json_obj, node_path)  # 获取json节点的值
        self.assertEqual(json_node_obj, expected_result)  # 断言是否相等


def set_code(class_type, function_name, code_str):
    """动态给指定的类添加方法的函数"""
    attr_method = None
    exec(code_str + '\nattr_method = %s' % function_name)
    setattr(class_type, function_name, attr_method)


def case1(method_name, row_value_list):
    """设定动态方法"""
    function_code_str = '''
def %(method_name)s(self):
    # 将行数据存储
    row_data = self.set_params(%(row_value_list)s)
    try:
        # 请求http协议接口
        resp = self.http_request(row_data[2], row_data[3], row_data[4])
        # json解析
        resp_json_obj = resp_util.json_decrypt(resp)
        # 断言结果
        self.check_resp(resp_json_obj, row_data[5], row_data[6], row_data[7])
    except Exception as e:
        raise e

    ''' % dict(
        method_name=method_name,
        row_value_list=row_value_list,
    )
    set_code(ExcelUnitTest, method_name, function_code_str)


def excel_driven(excel_file):
    """动态从Excel添加TestCase方法"""
    wb = load_workbook(filename=excel_file, read_only=True)
    ws = wb['Sheet1']
    # 设置测试套件
    test_suite = unittest.TestSuite()
    # 遍历Excel里内容
    for line_num, row in enumerate(ws.rows):
        if not isinstance(row[0], EmptyCell):  # 有遇到都是None的空行
            if line_num == 0:  # 第1行跳过
                continue
            row_value_list = map(lambda it: it.value, row)  # 读取每行数据
            temp_case_name = "test_%s" % (line_num)  # 设定运行时的testcase名称
            case1(temp_case_name, row_value_list)  # 处理数据
            test_suite.addTest(ExcelUnitTest(temp_case_name))  # 将case加入suite
    wb.close()
    # 返回测试套件
    return test_suite


def get_cell_values_by_cell_index(excel_file, sheet_index, cell_index):
    """从excel中通过行号获取该行所有数据"""
    wb = load_workbook(filename=excel_file, read_only=True)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[sheet_index])
    testcase_cn_name_list = map(lambda row: row[cell_index].value, ws.rows)
    del testcase_cn_name_list[0]  # 删除第1行的列名数据
    wb.close()
    return testcase_cn_name_list


def main():
    """主函数：执行excel文件，获取数据生成suite，运行unittest生成报告"""
    # 获取文件名
    excel_file = './unit_excel.xlsx'
    excel_file_name = os.path.basename(excel_file)

    # 从excel获取suite
    test_suite = excel_driven(excel_file)

    # 指定报告存储路径，运行
    report_save_path = './HTMLTestRunner.html'  # 确定生成报告的路径
    # 指定报告的case名称
    testcase_cn_name_list = get_cell_values_by_cell_index(excel_file, 0, 1)
    with open(report_save_path, 'wb') as fp:
        runner = HTMLTestRunner.HTMLTestRunner(
            verbosity=2,
            stream=fp,
            title='自动化测试报告',  # 配置报告信息
            description='详细测试用例结果',
            tester="RenYi",
            testclass_name=excel_file_name,
            testcass_list=testcase_cn_name_list
        )
        runner.run(test_suite)


if __name__ == '__main__':
    main()
